import json
import openpyxl
from copy import copy

# Configuration
input_file = 'DCF 1 CALLS- Savindi.xlsx'
output_file = 'reformatted_data.xlsx'

# Column Indices (0-based)
PREFIX_LEN = 70
BLOCK_SIZE = 32
NUM_BLOCKS = 15
# Within each block of 32, we only want the first 11 columns
KEEP_BLOCK_LEN = 11

SUFFIX_START_INDEX = 70 + (BLOCK_SIZE * NUM_BLOCKS)

def reformat_excel():
    print(f"Loading {input_file}...")
    try:
        wb = openpyxl.load_workbook(input_file)
        sheet = wb.active
    except Exception as e:
        print(f"Failed to load workbook: {e}")
        return

    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active
    new_sheet.title = "Reformatted Data"

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        print("No data found.")
        return

    original_header = rows[0]
    
    # Construct New Header
    # Prefix + Clean Block 1 Headers + Suffix
    prefix_header = list(original_header[:PREFIX_LEN])
    block_header = list(original_header[70:70+KEEP_BLOCK_LEN])
    suffix_header = list(original_header[SUFFIX_START_INDEX:])
    
    new_header = prefix_header + block_header + suffix_header
    new_sheet.append(new_header)

    print(f"Original Row Count: {len(rows)}")
    
    generated_rows = 0
    
    # Missing Data Collection
    missing_data_rows = []
    missing_data_header = ["Grama Niladhari ID", "Household ID", "Member ID", "Name", "Age", "Contact No", "All Members"]
    missing_data_rows.append(missing_data_header)
    
    COL_GN_ID = 5
    COL_HOUSEHOLD_ID = 47
    COL_CONTACT = 61

    for row_idx, row in enumerate(rows[1:], start=2): # Start from row 2
        prefix_data = list(row[:PREFIX_LEN])
        suffix_data = list(row[SUFFIX_START_INDEX:])
        
        gn_id = row[COL_GN_ID]
        household_id = row[COL_HOUSEHOLD_ID]
        
        # Priority: 10.1 (62), 10.2 (64), 10.3 (66), then 10 (61)
        contact_indices = [62, 64, 66, 61]
        contact_no = None
        for c_idx in contact_indices:
            val = row[c_idx]
            if val is not None and str(val).strip() != "":
                contact_no = val
                break
        
        # Determine number of members from Column 69 (Index 69)
        # Value might be string or int
        try:
            num_members_val = row[69]
            if num_members_val is None:
                num_members = 0
            else:
                num_members = int(float(num_members_val))
        except ValueError:
            print(f"Row {row_idx}: Could not parse member count '{row[69]}'. Defaulting to 0.")
            num_members = 0
            
        loop_count = max(1, num_members)
        
        # Pre-scan for All Members Summary
        all_members_list = []
        for k in range(loop_count):
            if k < NUM_BLOCKS:
                k_start = 70 + (k * BLOCK_SIZE)
                # Member Name index is +1 relative to block, Age is +2
                k_name = row[k_start + 1]
                k_age = row[k_start + 2]
                if k_name or k_age:
                     all_members_list.append({"name": str(k_name) if k_name else "", "age": str(k_age) if k_age else ""})
        
        all_members_json = json.dumps(all_members_list, ensure_ascii=False)
        
        for i in range(loop_count):
            # Always duplicate prefix/suffix for all member rows
            curr_prefix = prefix_data
            curr_suffix = suffix_data
                
            # Extract block data
            if i < NUM_BLOCKS:
                start = 70 + (i * BLOCK_SIZE)
                end_keep = start + KEEP_BLOCK_LEN
                block_data = list(row[start:end_keep])
                
                # Check for missing Name or Age
                # Relative indices: MemberID=0, Name=1, Age=2 (within block)
                m_id = block_data[0]
                m_name = block_data[1]
                m_age = block_data[2]
                
                # If Member ID is missing, ignore this block entirely (invalid missing data)
                if m_id is None or str(m_id).strip() == "":
                    # Still add to main sheet? Usually yes, generic empty row behavior.
                    # But for missing data report, we skip.
                    pass
                else:
                    is_name_missing = m_name is None or str(m_name).strip() == ""
                    is_age_missing = m_age is None or str(m_age).strip() == ""
                    
                    if is_name_missing or is_age_missing:
                        missing_data_rows.append([
                            gn_id,
                            household_id,
                            m_id,
                            m_name,
                            m_age,
                            contact_no,
                            all_members_json
                        ])
                    
            else:
                # If member count exceeds columns (unlikely), pad with None
                block_data = [None] * KEEP_BLOCK_LEN
            
            new_row = list(curr_prefix) + list(block_data) + list(curr_suffix)
            new_sheet.append(new_row)
            generated_rows += 1
        
    print(f"Generated {generated_rows} new rows.")
    new_wb.save(output_file)
    print(f"Saved to {output_file}")
    
    # Save Missing Data File if needed
    if len(missing_data_rows) > 1:
        print(f"Found {len(missing_data_rows)-1} members with missing data.")
        missing_wb = openpyxl.Workbook()
        missing_sheet = missing_wb.active
        missing_sheet.title = "Missing Data"
        
        for m_row in missing_data_rows:
            missing_sheet.append(m_row)
            
        missing_file = "missing_data.xlsx"
        missing_wb.save(missing_file)
        print(f"Saved missing data report to {missing_file}")
    else:
        print("No missing data found.")

if __name__ == "__main__":
    reformat_excel()
